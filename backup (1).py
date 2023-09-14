import concurrent.futures
import inspect
import time

import requests
from selenium.webdriver.common.by import By
from Driver import Driver
import openpyxl
import threading
import queue
import os

queue = queue.Queue()

lock = threading.Lock()


class Thread(threading.Thread):
    def __init__(self, t, *args):
        threading.Thread.__init__(self, target=t, args=args)
        self.start()


filename = "x.xlsx"


def write_to_excel(row, column, result):
    caller = inspect.getouterframes(inspect.currentframe())[1][3]
    print("Inside %s()" % caller)
    print("Acquiring lock")
    with lock:
        wookbook = openpyxl.load_workbook(filename)
        # Iterate the loop to read the cell values
        worksheet = wookbook.active
        print("Lock Acquired")
        worksheet.cell(row, column).value = result
        wookbook.save(filename)


def do_job(query, row):
    print(f'subbmitting row : {row}')
    driver = Driver()
    print('init success')
    # driver.get('https://www.google.com/')
    # time.sleep(0.3)
    driver.get(f'https://www.google.com/search?q={query}')
    driver.save_screenshot(f'screen{row}.png')

    cook = driver.safe_find('button', 'id', 'L2AGLb')
    if cook:
        'accepting cookies'
        cook.click()

    logo = driver.safe_find('div', 'class', 'logo')

    if not logo:
        print('No Logo')
        driver.quit()
        print('shutdown')
        return 'Shutdown', 0

    elem = driver.safe_find('div', 'jsname', 'TlEBqd')
    elem_num2 = driver.safe_find('div', 'class', 'liYKde g VjDLd')
    if elem or elem_num2:
        closed = driver.safe_find('div', 'jsname', 'RKTzrd')
        if closed:
            driver.quit()
            return 'Closed', row
        driver.quit()
        return 'Found', row

    else:
        driver.quit()
        return 'Not Found', row


def get_queries(start, end):
    wookbook = openpyxl.load_workbook(filename)
    # Iterate the loop to read the cell values
    worksheet = wookbook.active
    queries = []
    for i in range(start, end):
        counter = 1
        company = ''
        address = ''
        city = ''
        zip_code = ''
        for col in worksheet.iter_cols(2, 5):
            if counter == 1:
                company = col[i].value
                counter += 1
            elif counter == 2:
                address = col[i].value
                counter += 1

            elif counter == 3:
                city = col[i].value
                counter += 1

            else:
                zip_code = col[i].value
                counter = 1
        if city == None:
            city = ''
        if zip_code == None:
            zip_code = ''
        query = company.strip() + ' ' + address.strip() + str(city).strip() + str(zip_code).strip()
        query = query.replace(' ', '+')
        queries.append((query, i + 1))
    print("ba")
    return queries


def main(start, end):
    queries = get_queries(start, end)
    wookbook = openpyxl.load_workbook(filename)
    worksheet = wookbook.active
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:

        results = [executor.submit(do_job, query, row) for (query, row) in queries]
        try:
            for f in concurrent.futures.as_completed(results):
                res, row = f.result()
                if res == 'Shutdown':
                    raise 'Error'
                if res and row:
                    worksheet.cell(row, 9).value = res
                    print('saving', row)
        except:

            for i in results:
                i.cancel()
            executor.shutdown()

            print('No result canceld')

    wookbook.save('x.xlsx')


# 4930
if __name__ == '__main__':
    start = time.time()
    start_range = 9914
    end_range = 10000
    # driver = Driver()
    # driver.get("https://www.whatismyip-address.com/")
    # input()
    # driver.quit()
    main(start_range, end_range)
    # main()
    end = time.time()
    print(f"TOTAL Runtime of the program is {end - start}")

